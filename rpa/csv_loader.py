import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Mapping

from models.client_record import ClientRecord

EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
VALID_STATUS = {"ABERTO", "PAGO", "CANCELADO"}
EXPECTED_COLUMNS = [
    "id",
    "cliente_nome",
    "email",
    "status",
    "valor",
    "vencimento",
    "ultima_cobranca",
]
EXTRA_EXPORT_COLUMNS = ["enviado_em", "envio_status", "envio_erro"]
DATE_FIELDS = {"vencimento", "ultima_cobranca"}


def normalize_text(value: object | None) -> str | None:
    if value is None:
        return None

    normalized = value.strip() if isinstance(value, str) else str(value).strip()
    if not normalized:
        return None
    return normalized


def normalize_status(value: object | None) -> str | None:
    normalized = normalize_text(value)
    if normalized is None:
        return None
    return normalized.upper()


def normalize_date_input(value: object | None) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    normalized = normalize_text(value)
    if normalized is None:
        return None

    for date_format in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(normalized, date_format).date().isoformat()
        except ValueError:
            continue

    return normalized


def parse_valor(value: object | None) -> tuple[float | None, str | None]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = float(value)
        if parsed < 0:
            return None, "valor deve ser maior ou igual a zero"
        return parsed, None

    normalized = normalize_text(value)
    if normalized is None:
        return None, None

    candidate = normalized.replace(",", ".")
    try:
        parsed = float(candidate)
    except ValueError:
        return None, f"valor invalido '{normalized}'"

    if parsed < 0:
        return None, "valor deve ser maior ou igual a zero"
    return parsed, None


def is_valid_email(email: str | None) -> bool:
    normalized = normalize_text(email)
    if normalized is None:
        return False
    return bool(EMAIL_PATTERN.match(normalized))


def validate_record_fields(email: str | None, status: str | None, valor: float | None) -> list[str]:
    reasons: list[str] = []

    if not is_valid_email(email):
        reasons.append("email invalido")

    if status not in VALID_STATUS:
        reasons.append("status invalido (use ABERTO, PAGO ou CANCELADO)")

    if valor is not None and valor < 0:
        reasons.append("valor deve ser maior ou igual a zero")

    return reasons


def format_valor(valor: float | None) -> str:
    if valor is None:
        return ""
    return f"{valor:.2f}"


def _record_to_csv_row(record: ClientRecord, include_send_columns: bool) -> dict[str, str]:
    row: dict[str, str] = {
        "id": record.id or "",
        "cliente_nome": record.cliente_nome or "",
        "email": record.email or "",
        "status": record.status or "",
        "valor": format_valor(record.valor),
        "vencimento": record.vencimento or "",
        "ultima_cobranca": record.ultima_cobranca or "",
    }

    if include_send_columns:
        row["enviado_em"] = record.enviado_em or ""
        row["envio_status"] = record.envio_status or ""
        row["envio_erro"] = record.envio_erro or ""

    return row


def _validate_required_columns(fieldnames: list[str] | None, source_label: str) -> None:
    if fieldnames is None:
        raise ValueError(f"{source_label} sem cabecalho.")

    missing_columns = [column for column in EXPECTED_COLUMNS if column not in fieldnames]
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(f"{source_label} sem colunas obrigatorias: {missing}")


def _load_records_from_rows(
    rows: Iterable[tuple[int, Mapping[str, object | None]]],
) -> tuple[list[ClientRecord], list[str]]:
    records: list[ClientRecord] = []
    rejected_rows: list[str] = []
    used_ids: set[str] = set()
    next_auto_id = 1

    for line_number, row in rows:
        row_id = normalize_text(row.get("id"))
        cliente_nome = normalize_text(row.get("cliente_nome"))
        email = normalize_text(row.get("email"))
        status = normalize_status(row.get("status"))
        vencimento = normalize_date_input(row.get("vencimento"))
        ultima_cobranca = normalize_date_input(row.get("ultima_cobranca"))

        valor, valor_error = parse_valor(row.get("valor"))

        reasons: list[str] = []
        if row_id is None:
            while str(next_auto_id) in used_ids:
                next_auto_id += 1
            row_id = str(next_auto_id)
            used_ids.add(row_id)
            next_auto_id += 1
        else:
            used_ids.add(row_id)
            if row_id.isdigit():
                numeric_id = int(row_id)
                if numeric_id >= next_auto_id:
                    next_auto_id = numeric_id + 1

        reasons.extend(validate_record_fields(email, status, valor))
        if valor_error:
            reasons.append(valor_error)

        is_valid = len(reasons) == 0
        observacao_erro = "; ".join(reasons)
        selected = is_valid and status == "ABERTO"

        record = ClientRecord(
            id=row_id,
            cliente_nome=cliente_nome,
            email=email,
            status=status,
            valor=valor,
            vencimento=vencimento,
            ultima_cobranca=ultima_cobranca,
            observacao_erro=observacao_erro,
            is_valid=is_valid,
            selected=selected,
        )
        records.append(record)

        if not is_valid:
            rejected_rows.append(f"Linha {line_number}: {observacao_erro}")

    return records, rejected_rows


def _normalize_excel_field_value(field_name: str, value: object | None) -> object | None:
    if value is None:
        return None

    if field_name in DATE_FIELDS:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

    return value


def load_client_records(csv_path: str) -> tuple[list[ClientRecord], list[str]]:
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {csv_path}")

    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)

        _validate_required_columns(reader.fieldnames, source_label="CSV")

        rows = (
            (line_number, row)
            for line_number, row in enumerate(reader, start=1)
        )

        return _load_records_from_rows(rows)


def load_client_records_from_xlsx(
    xlsx_path: str,
    sheet_name: str | None = None,
) -> tuple[list[ClientRecord], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - protegido por requirements
        raise ImportError("Dependencia openpyxl nao encontrada. Instale openpyxl>=3.1.0.") from error

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {xlsx_path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Aba nao encontrada no XLSX: {sheet_name}")
            worksheet = workbook[sheet_name]

        rows_iter = worksheet.iter_rows(min_row=1, values_only=False)
        header_cells = next(rows_iter, None)
        if header_cells is None:
            raise ValueError("XLSX sem cabecalho.")

        fieldnames = [normalize_text(cell.value) or "" for cell in header_cells]
        _validate_required_columns(fieldnames, source_label="XLSX")

        def iter_rows() -> Iterable[tuple[int, dict[str, object | None]]]:
            for line_number, row_cells in enumerate(rows_iter, start=1):
                row_data: dict[str, object | None] = {}
                for index, field_name in enumerate(fieldnames):
                    if not field_name:
                        continue

                    value = row_cells[index].value if index < len(row_cells) else None
                    row_data[field_name] = _normalize_excel_field_value(field_name, value)

                yield line_number, row_data

        return _load_records_from_rows(iter_rows())
    finally:
        workbook.close()


def load_client_records_from_file(
    file_path: str,
    sheet_name: str | None = None,
) -> tuple[list[ClientRecord], list[str]]:
    extension = Path(file_path).suffix.lower()
    if extension == ".csv":
        return load_client_records(file_path)

    if extension == ".xlsx":
        return load_client_records_from_xlsx(file_path, sheet_name=sheet_name)

    raise ValueError("Formato de arquivo nao suportado. Use .csv ou .xlsx.")


def save_client_records(csv_path: str, records: list[ClientRecord], include_send_columns: bool = True) -> None:
    path = Path(csv_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [*EXPECTED_COLUMNS]
    if include_send_columns:
        fieldnames.extend(EXTRA_EXPORT_COLUMNS)

    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(_record_to_csv_row(record, include_send_columns=include_send_columns))
